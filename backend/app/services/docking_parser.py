"""
Parser khusus untuk file "REALISASI BIAYA DOCKING" (format laporan, bukan tabel flat).
Aturan ekstraksi (ditetapkan bareng user):
  1. Baris masuk Addendum kalau kolom Keterangan mengandung kata "tambahan".
  2. Baris masuk Induk kalau kolom Harga Utama > 0 - berlaku juga untuk status
     "pengembangan", "batal", "included", atau kosong.
  3. Uraian Pekerjaan digabung dari kolom setelah NO. sampai sebelum kolom Qty/Volume.
  4. Kategori Pekerjaan diambil dari baris header section (bertanda angka romawi di
     kolom NO.), dengan prefix angka romawi dibuang dari teksnya.
  5. Hanya baris dengan harga > 0 (bukan blank/"-") yang diekstrak.

Posisi kolom (Keterangan, Harga, dst) di-scan ulang tiap file karena tidak tetap.
"""
import io
import re

import openpyxl

ROMAN_TOKEN_RE = re.compile(r'^M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3})$', re.IGNORECASE)
YEAR_RE = re.compile(r'\b(20\d{2})\b')

NOISE_EXACT = {
    'realisasi', 'mulai', 'selesai', 'keterangan', 'kontrak induk', 'catatan',
    'total', 'tanda tangan', 'jabatan', 'nomor', 'tanggal', 'perihal', 'lampiran',
    'di t.tangani oleh',
}


def is_roman(s: str) -> bool:
    return bool(s) and bool(ROMAN_TOKEN_RE.match(s))


def norm(s) -> str:
    if s is None:
        return ""
    return str(s).strip().lower().rstrip(':').strip()


def norm_nospace(s) -> str:
    return re.sub(r'\s+', '', norm(s))


def strip_roman_prefix(text: str) -> str:
    m = re.match(r'^([IVXLCDM]+)([.\s]+)(.*)$', text.strip(), re.IGNORECASE)
    if m and is_roman(m.group(1)):
        return m.group(3).strip()
    return text.strip()


def find_header_idx(values):
    for i, row in enumerate(values):
        texts = [norm_nospace(c) for c in row if c]
        joined = ' '.join(texts)
        if 'uraian' in joined and any(k in joined for k in ('volume', 'qty', 'harga')):
            return i
    return None


def build_group_labels(header_row):
    labels = {}
    last = None
    for ci, val in enumerate(header_row):
        if val is not None and str(val).strip():
            last = str(val).strip()
        if last:
            labels[ci] = last
    return labels


def find_group_start(group_labels, must_have, must_not_have=(), after=None):
    for ci in sorted(group_labels.keys()):
        if after is not None and ci <= after:
            continue
        lbl = group_labels[ci].lower()
        if all(k in lbl for k in must_have) and not any(k in lbl for k in must_not_have):
            return ci
    return None


def find_harga_group_start(group_labels, must_have, must_not_have=(), after=None):
    candidates = []
    for ci in sorted(group_labels.keys()):
        if after is not None and ci <= after:
            continue
        lbl = group_labels[ci].lower()
        if all(k in lbl for k in must_have) and not any(k in lbl for k in must_not_have):
            candidates.append(ci)
    if not candidates:
        return None
    for ci in candidates:
        if 'rp' in group_labels[ci].lower():
            return ci
    return candidates[0]


def group_range(group_labels, group_start_col):
    if group_start_col is None:
        return None, None
    starts = sorted(set(group_labels.keys()))
    boundaries = []
    prev_label = None
    for ci in starts:
        if group_labels[ci] != prev_label:
            boundaries.append(ci)
            prev_label = group_labels[ci]
    boundaries.append(max(starts) + 1)
    for i, b in enumerate(boundaries):
        if b == group_start_col:
            return b, boundaries[i + 1]
    return group_start_col, group_start_col + 2


def max_num_in_range(row, start, end):
    best = None
    for ci in range(start, min(end, len(row))):
        v = row[ci]
        if v is None or (isinstance(v, str) and v.strip() in ('', '-')):
            continue
        try:
            n = float(v)
        except (TypeError, ValueError):
            continue
        if best is None or n > best:
            best = n
    return best


def col_for_sub(group_labels, sub_labels, group_start_col, want_sub):
    if group_start_col is None:
        return None
    target_label = group_labels[group_start_col]
    cols = [ci for ci, lbl in group_labels.items() if lbl == target_label]
    for ci in cols:
        if (sub_labels.get(ci) or '').strip().lower() == want_sub:
            return ci
    return cols[-1] if cols else None


def find_label_value(values, label_keys, max_row=25, max_scan=8):
    """Cari sel yang mengandung salah satu label (mis. 'nama kapal'), lalu ambil
    sel non-kosong pertama di sebelah kanannya (skip tanda ':')."""
    for row in values[:max_row]:
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            cell_norm = norm_nospace(cell)
            if any(key in cell_norm for key in label_keys):
                for cj in range(ci + 1, min(ci + max_scan, len(row))):
                    v = row[cj]
                    if v is not None and str(v).strip() not in (':', ''):
                        return str(v).strip()
    return ""


def guess_header(values, filename):
    nama_kapal = find_label_value(values, ['namakapal'])
    nama_perusahaan = find_label_value(values, ['pemilik'])
    tahun = find_label_value(values, ['periodedocking', 'dockingtahun', 'tahundocking'])
    if not tahun:
        m = YEAR_RE.search(filename)
        if m:
            tahun = m.group(1)
    return nama_kapal, nama_perusahaan, tahun


def parse_sheet(values, sheet_name):
    header_idx = find_header_idx(values)
    if header_idx is None:
        return [], [], [f"Header tabel item tidak ditemukan di sheet '{sheet_name}'"]

    header_row = values[header_idx]
    sub_row = values[header_idx + 1] if header_idx + 1 < len(values) else []
    group_labels = build_group_labels(header_row)
    sub_labels = {ci: (str(v).strip() if v else None) for ci, v in enumerate(sub_row)}

    uraian_col = find_group_start(group_labels, ['uraian'])
    if uraian_col is None:
        uraian_col = 1

    volume_col = find_group_start(group_labels, ['volume'])
    if volume_col is None:
        volume_col = find_group_start(group_labels, ['qty'])
    keterangan_col = find_group_start(group_labels, ['keterangan'])

    harga_utama_start = find_harga_group_start(group_labels, ['harga'], must_not_have=['batal', 'tambahan', 'realisasi'], after=volume_col or uraian_col)
    harga_utama_range = group_range(group_labels, harga_utama_start)

    tambahan_start = find_harga_group_start(group_labels, ['tambahan'])
    harga_tambahan_range = group_range(group_labels, tambahan_start)

    if volume_col is not None and group_labels.get(volume_col, '').strip().lower() == 'qty':
        qty_col = volume_col
        _, next_boundary = group_range(group_labels, volume_col)
        sat_col = next_boundary if group_labels.get(next_boundary, '').lower().startswith('sat') else None
    else:
        qty_col = col_for_sub(group_labels, sub_labels, volume_col, 'qty')
        sat_col = col_for_sub(group_labels, sub_labels, volume_col, 'sat')

    data_start = header_idx + 2
    end_col = volume_col if volume_col is not None else len(header_row)

    induk, addendum, warnings = [], [], []
    current_category = None
    seen_uraian_price_rows = {}

    for ri in range(data_start, len(values)):
        row = values[ri]
        if not any(c is not None for c in row):
            continue
        col_a = row[0]
        col_a_str = str(col_a).strip() if col_a is not None else ""

        if col_a_str and is_roman(col_a_str):
            cat_text = str(row[uraian_col]).strip() if len(row) > uraian_col and row[uraian_col] else ""
            cat_text = strip_roman_prefix(cat_text) or cat_text
            current_category = cat_text or current_category
            continue

        first_text = None
        for ci in range(0, min(uraian_col + 1, len(row))):
            if row[ci] is not None and str(row[ci]).strip():
                first_text = norm(row[ci])
                break
        if first_text and (first_text in NOISE_EXACT or any(first_text.startswith(k) for k in NOISE_EXACT)):
            continue

        frag = []
        for ci in range(uraian_col, end_col):
            if ci < len(row) and row[ci] is not None:
                v = str(row[ci]).strip()
                if v and v != '-':
                    frag.append(v)
        uraian_text = ' '.join(frag).strip()
        if not uraian_text:
            continue

        harga_utama_val = max_num_in_range(row, *harga_utama_range) if harga_utama_range[0] is not None else None
        harga_tambahan_val = max_num_in_range(row, *harga_tambahan_range) if harga_tambahan_range[0] is not None else None
        qty_val = row[qty_col] if qty_col is not None and qty_col < len(row) else None
        if isinstance(qty_val, float) and qty_val.is_integer():
            qty_val = int(qty_val)
        sat_val = row[sat_col] if sat_col is not None and sat_col < len(row) else None
        volume_satuan = f"{qty_val} {sat_val}".strip() if (qty_val or sat_val) else "-"

        keterangan_val = ""
        if keterangan_col is not None and keterangan_col < len(row) and row[keterangan_col]:
            keterangan_val = str(row[keterangan_col]).strip()

        is_tambahan = 'tambahan' in keterangan_val.lower()

        item = {
            'row': ri + 1,
            'kategori': current_category,
            'uraian': uraian_text,
            'volume_satuan': volume_satuan,
            'keterangan': keterangan_val,
        }

        if is_tambahan:
            if harga_tambahan_val and harga_tambahan_val > 0:
                item['harga'] = harga_tambahan_val
                addendum.append(item)
            elif harga_utama_val and harga_utama_val > 0:
                item['harga'] = harga_utama_val
                warnings.append(
                    f"Baris {ri+1}: keterangan 'tambahan' tapi harga cuma ada di kolom Harga Utama "
                    f"- masuk Addendum pakai harga utama, cek manual"
                )
                addendum.append(item)
        else:
            if harga_utama_val and harga_utama_val > 0:
                item['harga'] = harga_utama_val
                induk.append(item)
                seen_uraian_price_rows.setdefault(uraian_text, []).append(ri + 1)

    for uraian_text, rows_ in seen_uraian_price_rows.items():
        if len(rows_) > 1:
            warnings.append(
                f"Uraian '{uraian_text[:60]}' punya harga valid di lebih dari 1 baris: {rows_} "
                f"- cek manual, mungkin duplikat/baris lanjutan"
            )

    return induk, addendum, warnings


def _load_values(file_bytes: bytes, filename: str):
    name = filename.lower()
    if name.endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheetname = wb.sheet_names()[-1]
        sheet = wb.sheet_by_name(sheetname)
        values = [sheet.row_values(r) for r in range(min(sheet.nrows, 400))]
        values = [[(c if c != '' else None) for c in row] for row in values]
        return values, sheetname
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]
    values = list(ws.iter_rows(min_row=1, max_row=400, max_col=32, values_only=True))
    return values, sheetname


def parse_docking_file(file_bytes: bytes, filename: str) -> dict:
    values, sheetname = _load_values(file_bytes, filename)
    induk, addendum, warnings = parse_sheet(values, sheetname)
    nama_kapal, nama_perusahaan, tahun = guess_header(values, filename)
    return {
        "sheet_name": sheetname,
        "detected_nama_kapal": nama_kapal,
        "detected_nama_perusahaan": nama_perusahaan,
        "detected_tahun": tahun,
        "induk": induk,
        "addendum": addendum,
        "warnings": warnings,
    }
